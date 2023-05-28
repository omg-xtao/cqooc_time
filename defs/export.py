import shutil
from time import time
from typing import List, Union
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from model.exam import Exam
from model.task import Task

default_exam = Path("exams_init.xlsx")
result_exam = Path("exams.xlsx")


def export_exam_list(course_title: str, all_list: List[Union[Exam, Task]]) -> None:
    if not result_exam.is_file():
        shutil.copy(default_exam, result_exam)
    wb = load_workbook(result_exam)
    # 创建一个名为 课程名称 的 sheet 页
    ws = wb.create_sheet(course_title)
    if "Sheet" in wb.get_sheet_names():
        wb.remove_sheet(wb["Sheet"])
    # 标题
    ws["A1"] = "课程名称"
    ws["B1"] = "章节名称"
    ws["C1"] = "测验名称"
    ws["D1"] = "测验结束时间"
    ws["E1"] = "测验结束时间戳"
    ws["F1"] = "已回答答案"
    # 写入
    for i, data in enumerate(all_list):
        ws.cell(row=i + 2, column=1, value=data.course_title)
        ws.cell(row=i + 2, column=2, value=data.chapter)
        ws.cell(row=i + 2, column=3, value=data.title)
        ws.cell(row=i + 2, column=4, value=data.end_time)
        ws.cell(row=i + 2, column=5, value=data.submitEnd)
        ws.cell(row=i + 2, column=6, value=data.answer)
    # 将创建的工作簿保存为 exams.xlsx
    wb.save("exams.xlsx")
    # 最后关闭文件
    wb.close()


def export_exam_total(all_lists: List[Union[Exam, Task]]) -> None:
    wb: Workbook = load_workbook(result_exam)
    # 创建一个名为 课程名称 的 sheet 页
    ws = wb.get_sheet_by_name("total")
    # 标题
    ws["A1"] = "课程名称"
    ws["B1"] = "章节名称"
    ws["C1"] = "测验名称"
    ws["D1"] = "测验结束时间"
    ws["E1"] = "测验结束时间戳"
    ws["H1"] = int(time())
    # 写入
    for i, data in enumerate(all_lists):
        ws.cell(row=i + 2, column=1, value=data.course_title)
        ws.cell(row=i + 2, column=2, value=data.chapter)
        ws.cell(row=i + 2, column=3, value=data.title)
        ws.cell(row=i + 2, column=4, value=data.end_time)
        ws.cell(row=i + 2, column=5, value=data.submitEnd)
    # 将创建的工作簿保存为 exams.xlsx
    wb.save("exams.xlsx")
    # 最后关闭文件
    wb.close()
